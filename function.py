from openpyxl.styles import  PatternFill
from openpyxl.styles import Border,Side,colors

#实验组别对照的字母
group=['A','B','C','D','E','F','G','H','I','J','K']

#所有字段（指标）
title=['时间-天','AB','弹幕模块UV','弹幕开启UV','弹幕开启率','弹幕默认开启UV','弹幕默认开启率',
'弹幕手动开启UV','弹幕手动开启率','弹幕默认关闭UV','弹幕默认关闭率','弹幕手动关闭UV','弹幕手动关闭率',
'精简模式UV','精简模式开启率','默认精简模式UV','默认精简开启率','手动精简模式UV','手动精简开启率',
'弹幕发送UV','弹幕发送PV','弹幕发送量/弹幕模块UV','人均弹幕发送量','弹幕点赞UV','弹幕点赞PV','弹幕点赞量/弹幕模块UV',
'人均弹幕点赞量','弹幕点踩UV','弹幕点踩PV','弹幕点踩量/弹幕模块UV','人均点踩量','弹幕点赞比例','弹幕发布比例']
#不需要额外处理的字段
type1=['弹幕模块UV','弹幕开启UV','弹幕默认开启UV','弹幕手动开启UV','弹幕默认关闭UV','弹幕手动关闭UV',
    '精简模式UV','默认精简模式UV','手动精简模式UV','弹幕发送UV','弹幕发送PV','弹幕点赞UV','弹幕点赞PV',
    '弹幕点踩UV','弹幕点踩PV']
#需要额外处理的字段,这些参数需要更加细化的定量比较
type2=['弹幕开启率','弹幕默认开启率','弹幕手动开启率','弹幕默认关闭率','弹幕手动关闭率','精简模式开启率',
'默认精简开启率','手动精简开启率','弹幕发送量/弹幕模块UV','人均弹幕发送量','弹幕点赞量/弹幕模块UV',
'人均弹幕点赞量','弹幕点踩量/弹幕模块UV','人均点踩量','弹幕点赞比例','弹幕发布比例']

border_set = Border(left=Side(style='medium', color=colors.BLACK),
                    right=Side(style='medium', color=colors.BLACK),
                    top=Side(style='medium', color=colors.BLACK),
                    bottom=Side(style='medium', color=colors.BLACK))

#该函数确定取数的时间段&天数
def getBasicData(sheet,index):
    dateNum=1   # 取了dateNum天的数
    testNum=0   # 实验共testNum组
    date=[]     # 取数的时间段
    test=[]     # 组别名称
    colA=sheet[index[0]]
    colB=sheet[index[1]]
    tmp=''  
    #testNum计算 
    for item in colA:
        if tmp==item or tmp=='':
            testNum=testNum+1
        else:
            break
        tmp=item
    #test确定
    tmp=colB[1][:-1]
    for i in range(testNum):
        test.append(tmp+group[i])   
    #dateNum计算
    dateNum=int((len(sheet))/testNum)
    #date确定
    for i in range(dateNum):
        date.append(colA[testNum*i])
    return test,date
       
#该函数写入第一列（日期列）
def printCol_1(sheet,date,col_1):
    sheet['A2']=col_1
    for i in range(len(date)):
        sheet['A'+str(i+3)]=date[i]

#该函数写入后续数据
def printTable(sheet,df,date,test,colName):
    if colName=='弹幕模块UV':
        position=sheet.max_column
    else:
        position=sheet.max_column+1
    
    #不需要分别与AB组对比的情况
    if colName in type1:
        easyMode(sheet,df,date,test,colName,position)
    #需要分别与AB组对比的情况
    if colName in type2:
        hardMode(sheet,df,date,test,colName,position)              


def easyMode(sheet,df,date,test,colName,position):
    sheet.cell(row=1,column=position+1).value=colName
    for i in range(len(test)):
        cellItem=sheet.cell(row=2,column=position+i+1)
        cellItem.value=test[i]               #实验
        cellItem.border=border_set           #边框
        for j in range(len(date)):
            cellItem=sheet.cell(row=j+3,column=position+i+1)
            cellItem.value=str(df.at[(date[j],test[i]),colName])
            cellItem.border=border_set

def hardMode(sheet,df,date,test,colName,position):
    sheet.cell(row=1,column=position+1).value=colName
    #写入对照组A组
    sheet.cell(row=2,column=position+1).value=test[0]
    sheet.cell(row=2,column=position+1).border=border_set
    for j in range(len(date)):
        cellItem=sheet.cell(row=j+3,column=position+1)
        cellItem.value=str(df.at[(date[j],test[0]),colName])
        cellItem.border=border_set
    #写入实验组&对比数据
    for i in range(len(test)-2):
        sheet.cell(row=2,column=position+2+i).value=test[i+2]
        sheet.cell(row=2,column=position+2+i).border=border_set
        for j in range(len(date)):
            aCellValue=str(df.at[(date[j],test[0]),colName])
            cellValue=str(df.at[(date[j],test[i+2]),colName])
            sheet.cell(row=j+3,column=position+2+i).value=cellValue
            sheet.cell(row=j+3,column=position+2+i).border=border_set
            compare,cellColor=dataCompare(aCellValue,cellValue)
            cellNow=sheet.cell(row=len(date)+3+8+j,column=position+2+i)
            cellNow.value=compare
            cellNow.fill=PatternFill("solid", fgColor=cellColor)
    #重置position位置
    position=position+len(test)
    #写入对照组B组
    sheet.cell(row=2,column=position+1).value=test[1]
    sheet.cell(row=2,column=position+1).border=border_set
    for j in range(len(date)):
        cellItem=sheet.cell(row=j+3,column=position+1)
        cellItem.value=str(df.at[(date[j],test[1]),colName])
        cellItem.border=border_set
    #写入实验组
    for i in range(len(test)-2):
        sheet.cell(row=2,column=position+2+i).value=test[i+2]
        sheet.cell(row=2,column=position+2+i).border=border_set
        for j in range(len(date)):
            bCellValue=str(df.at[(date[j],test[1]),colName])
            cellValue=str(df.at[(date[j],test[i+2]),colName])
            sheet.cell(row=j+3,column=position+2+i).value=cellValue
            sheet.cell(row=j+3,column=position+2+i).border=border_set
            compare,cellColor=dataCompare(bCellValue,cellValue)
            cellNow=sheet.cell(row=len(date)+3+8+j,column=position+2+i)
            cellNow.value=compare
            cellNow.fill=PatternFill("solid", fgColor=cellColor)


def dataCompare(str1,str2):
    if str1[-1]=='%':
        predata=float(str2.strip('%'))-float(str1.strip('%'))
        result=str(round(predata,2))+'%'
        if result[0]=='-':
            cellColor='FFE2EFDA'
        else:
            cellColor='FFFDE2DF'
        return result,cellColor
    predata=float(str2)-float(str1)
    result=str(round(predata,4))
    if result[0]=='-':
        cellColor='FFF0FFF0'
    else:
        cellColor='FFFFF5EE'
    return result,cellColor

        


    
    


            

